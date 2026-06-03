#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
港股公告爬虫助手
从 Excel 持仓文件读取基金/账户/股票信息，爬取对应港股公告，
生成交互式财经日历网页，支持按基金/账户和股票筛选。

输出格式：
【股票代码】【股票名称】【日期】发布公告【公告名称】
"""

import json
import os
import re
import sys
from calendar import monthrange
from collections import defaultdict
from datetime import datetime, timedelta
from pathlib import Path

import requests

# API 配置
ANNOUNCEMENT_API = "https://np-anotice-stock.eastmoney.com/api/security/ann"
REQUEST_TIMEOUT = 30

# 项目路径
PROJECT_ROOT = Path(__file__).resolve().parent.parent
CONFIG_PATH = PROJECT_ROOT / "config" / "holdings.json"

# Excel 持仓文件路径（可通过 holdings.json 覆盖）
DEFAULT_EXCEL_PATH = (
    Path.home()
    / "OneDrive - TFI"
    / "From 若冲"
    / "@@@@@2026@@@@@"
    / "June"
    / "CA每日监测文件 rc.xlsx"
)


def load_holdings(excel_path: str | None = None) -> list[dict]:
    """
    读取持仓配置。优先从 Excel 读取，回退到 holdings.json。
    返回每条包含 code, name, fund, account。

    如果传入了 excel_path，读取后自动保存到 holdings.json。
    """
    # 情况1：命令行传入了 Excel 路径
    if excel_path:
        p = Path(excel_path)
        if not p.exists():
            print(f"[错误] 指定的 Excel 文件不存在: {excel_path}")
            sys.exit(1)
        holdings = _load_from_excel(p)
        # 保存到 holdings.json 供下次使用
        CONFIG_PATH.parent.mkdir(exist_ok=True)
        with open(CONFIG_PATH, "w", encoding="utf-8") as f:
            json.dump({"holdings": holdings, "excel_path": str(p)}, f, ensure_ascii=False, indent=2)
        print(f"[保存] 持仓已更新并保存到 {CONFIG_PATH}")
        return holdings

    # 情况2：先尝试默认 Excel 路径
    if DEFAULT_EXCEL_PATH.exists():
        holdings = _load_from_excel(DEFAULT_EXCEL_PATH)
        # 保存到 holdings.json
        CONFIG_PATH.parent.mkdir(exist_ok=True)
        with open(CONFIG_PATH, "w", encoding="utf-8") as f:
            json.dump({"holdings": holdings, "excel_path": str(DEFAULT_EXCEL_PATH)}, f, ensure_ascii=False, indent=2)
        return holdings

    # 情况3：回退到 holdings.json
    if CONFIG_PATH.exists():
        with open(CONFIG_PATH, "r", encoding="utf-8") as f:
            config = json.load(f)
        holdings = config.get("holdings", [])
        for h in holdings:
            h.setdefault("fund", "默认")
            h.setdefault("account", "")
        print(f"[JSON] 使用已保存的持仓数据（{len(holdings)} 条），如需更新请传入 --holdings")
        return holdings

    print(f"[错误] 未找到持仓数据源。请传入 --holdings 或创建以下之一：")
    print(f"  1. Excel 文件: {DEFAULT_EXCEL_PATH}")
    print(f"  2. JSON 配置: {CONFIG_PATH}")
    sys.exit(1)


def _load_from_excel(excel_path: Path) -> list[dict]:
    """从 Excel 读取持仓数据（第二个 sheet）。"""
    try:
        import openpyxl
    except ImportError:
        print("[错误] 缺少 openpyxl，请运行: pip install openpyxl")
        sys.exit(1)

    wb = openpyxl.load_workbook(str(excel_path), data_only=True)
    # 使用第二个 sheet（持仓汇总）
    ws = wb[wb.sheetnames[1]]
    rows = list(ws.iter_rows(values_only=True))

    if not rows or len(rows) < 2:
        print("[错误] Excel 文件为空或格式不正确")
        sys.exit(1)

    holdings = []
    seen = set()
    for row in rows[1:]:
        if row[0] is None or row[2] is None:
            continue
        fund = str(row[0]).strip().strip("'")
        account = str(row[1]).strip() if row[1] else ""
        code_raw = row[2]
        if isinstance(code_raw, (int, float)):
            code = str(int(code_raw)).zfill(5)
        else:
            code = str(code_raw).strip().zfill(5)
        name = str(row[3]).strip().strip("'") if row[3] else ""

        key = (fund, code)
        if key in seen:
            continue
        seen.add(key)

        holdings.append({
            "code": code,
            "name": name,
            "fund": fund,
            "account": account,
        })

    print(f"[Excel] 从 {excel_path.name} 读取到 {len(holdings)} 条持仓记录")
    funds = sorted(set(h["fund"] for h in holdings))
    print(f"[Excel] 涉及基金/账户: {', '.join(funds)}")
    return holdings


def fetch_announcements(stock_code: str, begin_time: str = None, end_time: str = None, days: int = None) -> list[dict]:
    """
    调用东方财富 API 获取指定港股代码的公告。
    """
    if begin_time and end_time:
        pass
    elif days:
        end_date = datetime.now()
        start_date = end_date - timedelta(days=days)
        end_time = end_date.strftime("%Y-%m-%d")
        begin_time = start_date.strftime("%Y-%m-%d")
    else:
        end_date = datetime.now()
        start_date = end_date - timedelta(days=7)
        end_time = end_date.strftime("%Y-%m-%d")
        begin_time = start_date.strftime("%Y-%m-%d")

    params = {
        "sr": "-1",
        "page_size": "100",
        "page_index": "1",
        "ann_type": "H",
        "client_source": "web",
        "stock_list": stock_code,
        "begin_time": begin_time,
        "end_time": end_time,
    }

    headers = {
        "User-Agent": (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/125.0.0.0 Safari/537.36"
        ),
        "Accept": "application/json, text/plain, */*",
        "Referer": "https://data.eastmoney.com/",
    }

    try:
        resp = requests.get(
            ANNOUNCEMENT_API,
            params=params,
            headers=headers,
            timeout=REQUEST_TIMEOUT,
        )
        resp.raise_for_status()
        data = resp.json()
    except requests.RequestException as e:
        print(f"  [网络错误] 获取 {stock_code} 公告失败: {e}")
        return []
    except json.JSONDecodeError as e:
        print(f"  [解析错误] 获取 {stock_code} 公告返回非法 JSON: {e}")
        return []

    if not data.get("data") or not data["data"].get("list"):
        return []

    announcements = []
    for item in data["data"]["list"]:
        codes = item.get("codes", [])
        if not codes:
            continue

        stock_info = codes[0]
        notice_date_raw = item.get("notice_date", "")
        notice_date = notice_date_raw.split()[0] if notice_date_raw else ""

        announcements.append({
            "art_code": item.get("art_code", ""),
            "stock_code": stock_info.get("stock_code", stock_code),
            "stock_name": stock_info.get("short_name", ""),
            "notice_date": notice_date,
            "title": item.get("title", "").strip(),
            "title_ch": item.get("title_ch", "").strip(),
            "columns": [col.get("column_name", "") for col in item.get("columns", [])],
        })

    return announcements


def format_output(announcements: list[dict], holdings_map: dict[str, str]) -> list[str]:
    """格式：【股票代码】【股票名称】【日期】发布公告【公告名称】"""
    lines = []
    for ann in announcements:
        code = ann["stock_code"]
        name = holdings_map.get(code, ann["stock_name"])
        date = ann["notice_date"]
        title = ann["title"] or ann["title_ch"]
        if not title:
            continue
        line = f"【{code}】【{name}】【{date}】发布公告【{title}】"
        lines.append(line)
    return lines


def main():
    """主入口。"""
    import argparse
    parser = argparse.ArgumentParser(description="港股公告爬虫")
    parser.add_argument("--holdings", type=str, default=None, help="持仓 Excel 文件路径，不传则使用已保存的数据")
    args = parser.parse_args()

    print("=" * 60)
    print("港股公告爬虫助手")
    print(f"运行时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("=" * 60)

    holdings = load_holdings(excel_path=args.holdings)
    holdings_map = {h["code"]: h.get("name", "") for h in holdings}

    # 读取已有公告数据（增量更新）
    data_file = PROJECT_ROOT / "data" / "announcements.json"
    existing_announcements: list[dict] = []
    if data_file.exists():
        with open(data_file, "r", encoding="utf-8") as f:
            existing_announcements = json.load(f)
        print(f"[增量] 已读取历史公告 {len(existing_announcements)} 条")

    # 计算数据获取范围：最近7天到未来一个月末
    today = datetime.now().date()
    begin_time = (today - timedelta(days=7)).strftime("%Y-%m-%d")

    # 自动补全：如果历史数据最早日期晚于今年4月1日，扩展范围补全
    april_first = today.replace(month=4, day=1).strftime("%Y-%m-%d")
    if existing_announcements:
        earliest = min(
            (d.get("notice_date", "") for d in existing_announcements if d.get("notice_date")),
            default="",
        )
        if earliest and earliest > april_first:
            begin_time = april_first
            print(f"\n[补全] 历史数据最早为 {earliest}，扩展查询范围至 {april_first}")

    if today.month == 12:
        next_month_first = today.replace(year=today.year + 1, month=1, day=1)
    else:
        next_month_first = today.replace(month=today.month + 1, day=1)
    if next_month_first.month == 12:
        next_next_month_first = next_month_first.replace(year=next_month_first.year + 1, month=1, day=1)
    else:
        next_next_month_first = next_month_first.replace(month=next_month_first.month + 1, day=1)
    end_time = (next_next_month_first - timedelta(days=1)).strftime("%Y-%m-%d")
    print(f"\n数据查询范围: {begin_time} ~ {end_time}")

    all_lines: list[str] = []
    all_announcements: list[dict] = []

    for holding in holdings:
        code = holding["code"]
        name = holding.get("name", code)
        fund = holding.get("fund", "")
        print(f"\n正在查询: [{fund}] {code} {name} ...")

        announcements = fetch_announcements(code, begin_time=begin_time, end_time=end_time)
        if not announcements:
            print(f"  该时间段暂无公告")
            continue

        print(f"  获取到 {len(announcements)} 条公告")
        # 为每条公告附加 fund 信息
        for ann in announcements:
            ann["fund"] = fund
            ann["account"] = holding.get("account", "")

        lines = format_output(announcements, holdings_map)
        all_lines.extend(lines)
        all_announcements.extend(announcements)

    print("\n" + "=" * 60)
    print("查询结果汇总")
    print("=" * 60)

    # 合并新旧数据（按 art_code 去重）
    seen_art_codes = set()
    merged_announcements = []
    for ann in existing_announcements + all_announcements:
        art_code = ann.get("art_code", "")
        key = art_code if art_code else f"{ann['stock_code']}_{ann.get('notice_date', '')}_{ann.get('title', '')}"
        if key not in seen_art_codes:
            seen_art_codes.add(key)
            merged_announcements.append(ann)

    print(f"\n历史公告: {len(existing_announcements)} 条")
    print(f"本次新增: {len(all_announcements)} 条")
    print(f"合并后共: {len(merged_announcements)} 条")

    # 保存合并后的数据
    data_file.parent.mkdir(exist_ok=True)
    with open(data_file, "w", encoding="utf-8") as f:
        json.dump(merged_announcements, f, ensure_ascii=False, indent=2)
    print(f"[保存] 合并数据已保存到 {data_file}")

    if not all_lines:
        print("\n该时间段所有持仓股票均暂无新公告。\n")
        _generate_calendar_html(merged_announcements, holdings)
        return

    for line in all_lines:
        print(line)

    # 将结果写入日志文件
    log_dir = PROJECT_ROOT / "logs"
    log_dir.mkdir(exist_ok=True)
    log_file = log_dir / f"announcements_{datetime.now().strftime('%Y%m%d')}.txt"

    with open(log_file, "w", encoding="utf-8") as f:
        f.write(f"港股公告汇总 - {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
        f.write("=" * 60 + "\n")
        for line in all_lines:
            f.write(line + "\n")

    print(f"\n结果已保存至: {log_file}")
    print("=" * 60)

    # 生成财经日历 HTML
    _generate_calendar_html(merged_announcements, holdings)


def _generate_calendar_html(announcements: list[dict], holdings: list[dict]) -> None:
    """生成完整月历视图的 HTML 页面，含基金/账户筛选。"""
    # 按日期分组
    date_map: dict[str, list[dict]] = defaultdict(list)
    for ann in announcements:
        date_map[ann["notice_date"]].append(ann)

    # 为每个股票分配一个稳定颜色
    stock_colors = [
        "#3b82f6", "#ef4444", "#10b981", "#f59e0b",
        "#8b5cf6", "#ec4899", "#06b6d4", "#f97316",
        "#84cc16", "#6366f1", "#14b8a6", "#d946ef",
        "#f43f5e", "#0ea5e9", "#a855f7", "#22c55e",
    ]
    code_to_color = {}
    color_idx = 0
    for ann in announcements:
        code = ann["stock_code"]
        if code not in code_to_color:
            code_to_color[code] = stock_colors[color_idx % len(stock_colors)]
            color_idx += 1

    # 为每个基金分配颜色
    fund_colors = [
        "#0ea5e9", "#f97316", "#8b5cf6", "#ec4899", "#10b981",
        "#ef4444", "#3b82f6", "#f59e0b", "#06b6d4", "#84cc16",
    ]
    funds = sorted(set(h["fund"] for h in holdings))
    fund_to_color = {}
    for i, fund in enumerate(funds):
        fund_to_color[fund] = fund_colors[i % len(fund_colors)]

    # 提取中文名的正则
    import re as _re
    _cn_re = _re.compile(r'^[\u4e00-\u9fa5]+')

    # 公告类型映射（从东方财富 column_name 映射到用户分类）
    CATEGORY_MAP = {
        "股东大会": ["股東周年大會通告", "股東特別大會通告", "股東周年大會的結果", "股東特別大會的結果", "委任代表表格"],
        "业绩快报": ["季度業績", "資產淨值", "月報表"],
        "交易披露": ["股份購回", "根據一般性授權發行股份", "發行可轉換證券", "回購股份的說明函件", "更改證券條款或隨附於證券的權利"],
        "中期业绩": ["季度業績"],
        "末期业绩": ["季度業績"],
        "重大事项": ["發出通函後的重大資料", "更換董事或重要行政職能或職責的變更", "更換審核委員會成員", "更換提名委員會成員", "更換薪酬委員會成員", "董事會召開日期", "董事名單和他們的地位和作用", "憲章文件", "修訂憲章文件", "在股東批准的情況下重選或委任董事", "暫停辦理過戶登記手續或更改暫停辦理過戶日期"],
        "关联交易": ["海外監管公告-證券發行及相關事宜", "海外監管公告-業務發展最新情況"],
        "会议表决": ["股東周年大會的結果", "股東特別大會的結果", "委任代表表格"],
    }
    # 反向映射：column_name -> set of categories
    _col_to_cats = {}
    for cat, cols in CATEGORY_MAP.items():
        for col in cols:
            _col_to_cats.setdefault(col, set()).add(cat)

    # 准备 JS 数据
    js_date_map = {}
    for date_str, anns in date_map.items():
        js_date_map[date_str] = []
        for ann in anns:
            code = ann["stock_code"]
            name = ann.get("stock_name", "")
            # 优先用 holdings 里的名称
            for h in holdings:
                if h["code"] == code:
                    name = h["name"]
                    break
            title = ann["title"] or ann["title_ch"]
            # 提取中文名
            cn_match = _cn_re.match(name)
            name_cn = cn_match.group(0) if cn_match else name
            art_code = ann.get("art_code", "")
            url = f"https://data.eastmoney.com/notices/detail/{art_code}/{code}.html" if art_code else ""
            pdf_url = f"https://pdf.dfcfw.com/pdf/H2_{art_code}_1.pdf" if art_code else ""
            # 计算该公告所属的分类
            cats = set()
            for col in ann.get("columns", []):
                for c in _col_to_cats.get(col, []):
                    cats.add(c)
            js_date_map[date_str].append({
                "code": code,
                "name": name,
                "name_cn": name_cn,
                "title": title,
                "color": code_to_color.get(code, "#64748b"),
                "fund": ann.get("fund", ""),
                "fundColor": fund_to_color.get(ann.get("fund", ""), "#64748b"),
                "url": url,
                "pdf_url": pdf_url,
                "art_code": art_code,
                "categories": sorted(cats),
            })

    stocks = []
    seen_codes = set()
    for h in holdings:
        code = h["code"]
        if code not in seen_codes:
            seen_codes.add(code)
            full_name = h.get("name", "")
            cn_match = _cn_re.match(full_name)
            name_cn = cn_match.group(0) if cn_match else full_name
            stocks.append({
                "code": code,
                "name": full_name,
                "name_cn": name_cn,
                "color": code_to_color.get(code, "#64748b"),
                "fund": h.get("fund", ""),
                "account": h.get("account", ""),
            })

    # 按股票代码数字排序
    stocks.sort(key=lambda s: int(s["code"]))

    # 基金列表
    funds_list = [{"name": f, "color": fund_to_color[f]} for f in funds]

    # 生成每条公告的独立详情页
    notices_dir = PROJECT_ROOT / "notices"
    notices_dir.mkdir(exist_ok=True)
    for ann in announcements:
        art_code = ann.get("art_code", "")
        code = ann["stock_code"]
        if not art_code:
            continue
        detail_path = notices_dir / f"{art_code}_{code}.html"
        detail_html = _build_notice_detail(ann, code_to_color, fund_to_color, _cn_re)
        with open(detail_path, "w", encoding="utf-8") as f:
            f.write(detail_html)

    # 更新 js_date_map 中的 url 为本地详情页路径
    for date_str, anns in js_date_map.items():
        for a in anns:
            a["url"] = f"./notices/{a.get('art_code', '')}_{a['code']}.html"

    # 收集所有实际出现的分类并分配颜色
    cat_colors = {
        "股东大会": "#ef4444",
        "业绩快报": "#3b82f6",
        "交易披露": "#10b981",
        "中期业绩": "#f59e0b",
        "末期业绩": "#8b5cf6",
        "重大事项": "#f97316",
        "关联交易": "#ec4899",
        "会议表决": "#06b6d4",
    }
    seen_cats = set()
    for anns in js_date_map.values():
        for a in anns:
            seen_cats.update(a.get("categories", []))
    categories_list = [{"name": c, "color": cat_colors.get(c, "#64748b")} for c in sorted(seen_cats)]

    data_json = json.dumps(js_date_map, ensure_ascii=False)
    stocks_json = json.dumps(stocks, ensure_ascii=False)
    funds_json = json.dumps(funds_list, ensure_ascii=False)
    categories_json = json.dumps(categories_list, ensure_ascii=False)
    today_str = datetime.now().strftime("%Y-%m-%d")
    update_time = datetime.now().strftime("%Y-%m-%d %H:%M")

    html_path = PROJECT_ROOT / "index.html"
    html_content = _build_html(data_json, stocks_json, funds_json, categories_json, today_str, update_time)

    with open(html_path, "w", encoding="utf-8") as f:
        f.write(html_content)

    print(f"\n财经日历已生成: {html_path}")
    print(f"公告详情页已生成: {notices_dir}")
    print("可直接用浏览器打开查看。")


def _build_notice_detail(ann: dict, code_to_color: dict, fund_to_color: dict, cn_re) -> str:
    """为单条公告生成详情页 HTML。"""
    code = ann["stock_code"]
    name = ann.get("stock_name", "")
    cn_match = cn_re.match(name)
    name_cn = cn_match.group(0) if cn_match else name
    title = ann["title"] or ann["title_ch"]
    notice_date = ann.get("notice_date", "")
    fund = ann.get("fund", "")
    color = code_to_color.get(code, "#64748b")
    fund_color = fund_to_color.get(fund, "#64748b")
    art_code = ann.get("art_code", "")

    # 外部搜索链接
    hkex_search = f"https://www1.hkexnews.hk/search/titlesearch.xhtml?lang=zh&stockCode={code.zfill(5)}"
    sina_url = f"https://stock.finance.sina.com.cn/hkstock/notice/{code}.html"
    eastmoney_search = f"https://data.eastmoney.com/notices/hk/{code}.html"
    pdf_url = f"https://pdf.dfcfw.com/pdf/H2_{art_code}_1.pdf" if art_code else ""

    return f'''<!DOCTYPE html>
<html lang="zh-CN">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>{title}</title>
    <style>
        * {{ margin: 0; padding: 0; box-sizing: border-box; }}
        body {{
            font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, "Helvetica Neue", Arial, sans-serif;
            background: linear-gradient(135deg, #f0f4f8 0%, #e8eef5 100%);
            min-height: 100vh;
            color: #1e293b;
            padding: 24px 16px;
        }}
        .container {{
            max-width: 800px;
            margin: 0 auto;
        }}
        .back-link {{
            display: inline-block;
            margin-bottom: 16px;
            color: #3b82f6;
            text-decoration: none;
            font-size: 14px;
            font-weight: 500;
        }}
        .back-link:hover {{ text-decoration: underline; }}
        .card {{
            background: #fff;
            border-radius: 16px;
            box-shadow: 0 1px 3px rgba(0,0,0,0.06), 0 1px 2px rgba(0,0,0,0.04);
            padding: 28px;
            margin-bottom: 16px;
        }}
        .fund-tag {{
            display: inline-block;
            font-size: 12px;
            font-weight: 700;
            padding: 4px 10px;
            border-radius: 6px;
            margin-bottom: 12px;
            letter-spacing: 0.3px;
            background-color: {fund_color}15;
            color: {fund_color};
        }}
        .stock-info {{
            display: flex;
            align-items: center;
            gap: 8px;
            margin-bottom: 16px;
        }}
        .stock-code {{
            font-family: "SF Mono", Monaco, monospace;
            font-size: 14px;
            font-weight: 700;
            padding: 4px 10px;
            border-radius: 6px;
            background-color: {color}15;
            color: {color};
            border: 1px solid {color}30;
        }}
        .stock-name {{
            font-size: 16px;
            font-weight: 600;
            color: #0f172a;
        }}
        .title {{
            font-size: 22px;
            font-weight: 700;
            color: #0f172a;
            line-height: 1.4;
            margin-bottom: 16px;
        }}
        .meta {{
            display: flex;
            flex-wrap: wrap;
            gap: 12px;
            margin-bottom: 24px;
            padding-bottom: 20px;
            border-bottom: 1px solid #f1f5f9;
        }}
        .meta-item {{
            font-size: 13px;
            color: #64748b;
        }}
        .meta-label {{ color: #94a3b8; }}
        .section-title {{
            font-size: 14px;
            font-weight: 600;
            color: #0f172a;
            margin-bottom: 12px;
        }}
        .link-list {{
            display: flex;
            flex-direction: column;
            gap: 8px;
        }}
        .link-item {{
            display: flex;
            align-items: center;
            gap: 8px;
            padding: 10px 14px;
            background: #f8fafc;
            border-radius: 8px;
            text-decoration: none;
            color: #334155;
            font-size: 14px;
            transition: background 0.15s;
        }}
        .link-item:hover {{ background: #f1f5f9; }}
        .link-arrow {{
            margin-left: auto;
            color: #94a3b8;
            font-size: 12px;
        }}
        .notice {{
            font-size: 13px;
            color: #94a3b8;
            margin-top: 20px;
            padding: 12px;
            background: #f8fafc;
            border-radius: 8px;
        }}
    </style>
</head>
<body>
    <div class="container">
        <a href="../index.html" class="back-link">← 返回日历</a>
        <div class="card">
            <div class="fund-tag">{fund}</div>
            <div class="stock-info">
                <span class="stock-code">{code}</span>
                <span class="stock-name">{name_cn}</span>
            </div>
            <div class="title">{title}</div>
            <div class="meta">
                <div class="meta-item"><span class="meta-label">公告日期:</span> {notice_date}</div>
                <div class="meta-item"><span class="meta-label">股票代码:</span> {code}</div>
                <div class="meta-item"><span class="meta-label">所属基金:</span> {fund}</div>
            </div>
            <div class="section-title">查看原始公告</div>
            <div class="link-list">
                <a class="link-item" href="{pdf_url}" target="_blank">
                    <span>📄 东方财富 PDF 直链</span>
                    <span class="link-arrow">↗</span>
                </a>
                <a class="link-item" href="{eastmoney_search}" target="_blank">
                    <span>东方财富 - {name_cn} 公告列表</span>
                    <span class="link-arrow">↗</span>
                </a>
                <a class="link-item" href="{sina_url}" target="_blank">
                    <span>新浪财经 - {name_cn} 公告列表</span>
                    <span class="link-arrow">↗</span>
                </a>
                <a class="link-item" href="{hkex_search}" target="_blank">
                    <span>港交所披露易 - 标题搜索</span>
                    <span class="link-arrow">↗</span>
                </a>
            </div>
            <div class="notice">
                提示：若上方 PDF 直链无法打开，可通过其他平台链接查看原始公告。
            </div>
        </div>
    </div>
</body>
</html>'''


def _build_html(data_json: str, stocks_json: str, funds_json: str, categories_json: str, today_str: str, update_time: str) -> str:
    """构建完整月历视图的 HTML，含年月下拉选择器、基金筛选、股票筛选。"""

    template = '''<!DOCTYPE html>
<html lang="zh-CN">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>港股公告日历</title>
    <style>
        * {{ margin: 0; padding: 0; box-sizing: border-box; }}
        body {{
            font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, "Helvetica Neue", Arial, sans-serif;
            background: linear-gradient(135deg, #f0f4f8 0%, #e8eef5 100%);
            min-height: 100vh;
            color: #1e293b;
        }}
        .container {{
            max-width: 1400px;
            margin: 0 auto;
            padding: 24px 16px;
        }}
        header {{
            text-align: center;
            margin-bottom: 16px;
        }}
        header h1 {{
            font-size: 26px;
            font-weight: 700;
            color: #0f172a;
            letter-spacing: -0.5px;
        }}
        header .update-time {{
            font-size: 12px;
            color: #94a3b8;
            margin-top: 4px;
        }}
        .holdings-btn {{
            background: #3b82f6;
            color: #fff;
            border: none;
            border-radius: 8px;
            padding: 7px 14px;
            font-size: 13px;
            font-weight: 500;
            cursor: pointer;
            transition: background 0.15s;
        }}
        .holdings-btn:hover {{ background: #2563eb; }}
        /* 持仓管理模态框 */
        .holdings-overlay {{
            position: fixed;
            inset: 0;
            background: rgba(0,0,0,0.45);
            z-index: 2000;
            display: none;
            align-items: center;
            justify-content: center;
        }}
        .holdings-overlay.active {{ display: flex; }}
        .holdings-modal {{
            background: #fff;
            border-radius: 16px;
            width: 900px;
            max-width: 95vw;
            max-height: 90vh;
            display: flex;
            flex-direction: column;
            box-shadow: 0 25px 50px -12px rgba(0,0,0,0.25);
        }}
        .holdings-header {{
            display: flex;
            align-items: center;
            justify-content: space-between;
            padding: 16px 20px;
            border-bottom: 1px solid #e2e8f0;
        }}
        .holdings-title {{ font-size: 18px; font-weight: 700; color: #0f172a; }}
        .holdings-close {{
            background: none; border: none; font-size: 24px; color: #94a3b8;
            cursor: pointer; line-height: 1;
        }}
        .holdings-body {{
            padding: 16px 20px;
            overflow-y: auto;
            flex: 1;
        }}
        .holdings-actions {{
            display: flex;
            gap: 10px;
            padding: 12px 20px;
            border-top: 1px solid #e2e8f0;
            background: #f8fafc;
            border-radius: 0 0 16px 16px;
        }}
        .holdings-actions button {{
            padding: 8px 16px;
            border-radius: 8px;
            border: none;
            font-size: 13px;
            font-weight: 500;
            cursor: pointer;
        }}
        .holdings-actions .btn-primary {{ background: #3b82f6; color: #fff; }}
        .holdings-actions .btn-primary:hover {{ background: #2563eb; }}
        .holdings-actions .btn-secondary {{ background: #e2e8f0; color: #334155; }}
        .holdings-actions .btn-secondary:hover {{ background: #cbd5e1; }}
        .holdings-group {{
            margin-bottom: 16px;
        }}
        .holdings-group-title {{
            font-size: 14px;
            font-weight: 600;
            color: #0f172a;
            margin-bottom: 8px;
            padding: 6px 10px;
            background: #f1f5f9;
            border-radius: 6px;
            display: flex;
            align-items: center;
            gap: 6px;
        }}
        .holdings-group-dot {{
            width: 10px; height: 10px; border-radius: 50%;
        }}
        .holdings-table {{
            width: 100%;
            border-collapse: collapse;
            font-size: 13px;
        }}
        .holdings-table th {{
            text-align: left;
            padding: 8px 10px;
            color: #64748b;
            font-weight: 500;
            border-bottom: 1px solid #e2e8f0;
        }}
        .holdings-table td {{
            padding: 7px 10px;
            border-bottom: 1px solid #f1f5f9;
            color: #334155;
        }}
        .holdings-table tr:hover td {{ background: #f8fafc; }}
        .holdings-upload-area {{
            border: 2px dashed #cbd5e1;
            border-radius: 10px;
            padding: 24px;
            text-align: center;
            color: #64748b;
            margin-bottom: 16px;
            cursor: pointer;
            transition: border-color 0.15s;
        }}
        .holdings-upload-area:hover {{ border-color: #3b82f6; }}
        .holdings-upload-area.dragover {{ border-color: #3b82f6; background: #eff6ff; }}
        .holdings-preview {{
            background: #f8fafc;
            border-radius: 10px;
            padding: 12px;
            margin-bottom: 16px;
            max-height: 300px;
            overflow-y: auto;
        }}
        .holdings-tab-bar {{
            display: flex;
            gap: 4px;
            margin-bottom: 12px;
        }}
        .holdings-tab {{
            padding: 6px 14px;
            border-radius: 6px;
            font-size: 13px;
            font-weight: 500;
            cursor: pointer;
            border: none;
            background: #e2e8f0;
            color: #64748b;
        }}
        .holdings-tab.active {{ background: #3b82f6; color: #fff; }}
        /* 控制栏 */
        .control-bar {{
            display: flex;
            align-items: center;
            justify-content: center;
            flex-wrap: wrap;
            gap: 12px;
            margin-bottom: 12px;
            background: #fff;
            padding: 12px 16px;
            border-radius: 12px;
            box-shadow: 0 1px 3px rgba(0,0,0,0.04);
        }}
        .control-group {{
            display: flex;
            align-items: center;
            gap: 6px;
        }}
        .control-label {{
            font-size: 13px;
            color: #64748b;
            font-weight: 500;
        }}
        .control-select {{
            padding: 7px 10px;
            border: 1px solid #e2e8f0;
            border-radius: 8px;
            font-size: 14px;
            color: #0f172a;
            background: #fff;
            cursor: pointer;
            outline: none;
            transition: border-color 0.15s;
        }}
        .control-select:hover {{
            border-color: #cbd5e1;
        }}
        .control-select:focus {{
            border-color: #3b82f6;
        }}
        .today-btn {{
            padding: 7px 14px;
            background: #3b82f6;
            color: #fff;
            border: none;
            border-radius: 8px;
            font-size: 13px;
            font-weight: 600;
            cursor: pointer;
            transition: background 0.15s;
        }}
        .today-btn:hover {{
            background: #2563eb;
        }}
        /* 筛选栏 */
        .filter-section {{
            display: flex;
            flex-direction: column;
            gap: 8px;
            margin-bottom: 12px;
        }}
        .filter-bar {{
            display: flex;
            align-items: center;
            justify-content: center;
            flex-wrap: wrap;
            gap: 8px 14px;
            background: #fff;
            padding: 10px 16px;
            border-radius: 12px;
            box-shadow: 0 1px 3px rgba(0,0,0,0.04);
        }}
        .filter-label {{
            font-size: 13px;
            color: #64748b;
            font-weight: 500;
            margin-right: 4px;
        }}
        .filter-item {{
            display: flex;
            align-items: center;
            gap: 5px;
            cursor: pointer;
            padding: 4px 8px;
            border-radius: 6px;
            transition: background 0.12s;
            user-select: none;
        }}
        .filter-item:hover {{
            background: #f1f5f9;
        }}
        .filter-item input[type="checkbox"] {{
            width: 15px;
            height: 15px;
            cursor: pointer;
            accent-color: #3b82f6;
        }}
        .filter-dot {{
            width: 10px;
            height: 10px;
            border-radius: 50%;
            flex-shrink: 0;
        }}
        .filter-text {{
            font-size: 13px;
            color: #334155;
            font-weight: 500;
        }}
        /* 股票筛选栏 - 7列网格 */
        #stockFilterBar {{
            display: grid;
            grid-template-columns: repeat(7, 1fr);
            gap: 4px 8px;
            align-items: center;
        }}
        #stockFilterBar .filter-label {{
            grid-column: 1 / -1;
            margin-right: 0;
            margin-bottom: 4px;
        }}
        #stockFilterBar .filter-item {{
            padding: 3px 6px;
            justify-content: flex-start;
            overflow: hidden;
        }}
        #stockFilterBar .filter-text {{
            display: flex;
            align-items: center;
            min-width: 0;
        }}
        #stockFilterBar .stock-name {{
            overflow: hidden;
            text-overflow: ellipsis;
            white-space: nowrap;
        }}
        #stockFilterBar .filter-actions {{
            grid-column: 1 / -1;
            margin-top: 4px;
            margin-left: 0;
            padding-left: 0;
            border-left: none;
            justify-self: start;
        }}
        .stock-code {{
            display: inline-block;
            width: 48px;
            font-family: "SF Mono", Monaco, "Cascadia Code", "Roboto Mono", Consolas, "Courier New", monospace;
            font-variant-numeric: tabular-nums;
            font-weight: 600;
            color: #1e293b;
            margin-right: 4px;
        }}
        .stock-name {{
            color: #475569;
        }}
        .filter-actions {{
            display: flex;
            gap: 6px;
            margin-left: 8px;
            padding-left: 12px;
            border-left: 1px solid #e2e8f0;
        }}
        .filter-action-btn {{
            font-size: 12px;
            color: #3b82f6;
            background: none;
            border: none;
            cursor: pointer;
            padding: 3px 6px;
            border-radius: 4px;
            font-weight: 500;
        }}
        .filter-action-btn:hover {{
            background: #eff6ff;
        }}
        /* 星期标题 */
        .weekday-header {{
            display: grid;
            grid-template-columns: repeat(7, 1fr);
            gap: 8px;
            margin-bottom: 8px;
        }}
        .weekday-cell {{
            text-align: center;
            font-size: 13px;
            font-weight: 600;
            color: #64748b;
            padding: 8px 0;
        }}
        .calendar-grid {{
            display: grid;
            grid-template-columns: repeat(7, 1fr);
            grid-template-rows: repeat(6, minmax(150px, auto));
            gap: 8px;
        }}
        .day-cell {{
            background: #fff;
            border-radius: 10px;
            box-shadow: 0 1px 2px rgba(0,0,0,0.04);
            overflow: hidden;
            display: flex;
            flex-direction: column;
            transition: transform 0.1s, box-shadow 0.1s;
            position: relative;
        }}
        .day-cell:hover {{
            transform: translateY(-1px);
            box-shadow: 0 4px 12px rgba(0,0,0,0.06);
        }}
        .day-cell.today {{
            box-shadow: 0 0 0 2px #3b82f6, 0 2px 8px rgba(59,130,246,0.12);
        }}
        .day-cell.other-month {{
            background: #f8fafc;
        }}
        .day-cell.other-month .day-number {{
            color: #cbd5e1;
        }}
        .day-header-cell {{
            padding: 8px 10px 4px;
            display: flex;
            justify-content: space-between;
            align-items: center;
        }}
        .day-number {{
            font-size: 16px;
            font-weight: 700;
            color: #0f172a;
        }}
        .day-cell.today .day-number {{
            color: #3b82f6;
        }}
        .day-cell.other-month .day-number {{
            color: #cbd5e1;
        }}
        .ann-count {{
            font-size: 11px;
            color: #fff;
            background: #3b82f6;
            padding: 1px 6px;
            border-radius: 10px;
            font-weight: 600;
        }}
        .day-body {{
            padding: 0 8px 8px;
            flex: 1;
            display: flex;
            flex-direction: column;
            gap: 5px;
            overflow-y: auto;
            max-height: 220px;
        }}
        .ann-card {{
            background: #f8fafc;
            border-radius: 6px;
            padding: 6px 8px;
            cursor: pointer;
            transition: background 0.12s;
            border-left: 3px solid transparent;
        }}
        .ann-card:hover {{
            background: #f1f5f9;
        }}
        .stock-tag {{
            display: inline-flex;
            align-items: center;
            gap: 4px;
            font-size: 10px;
            font-weight: 600;
            margin-bottom: 2px;
        }}
        .fund-tag {{
            display: inline-block;
            font-size: 9px;
            font-weight: 700;
            padding: 1px 5px;
            border-radius: 4px;
            margin-bottom: 3px;
            letter-spacing: 0.3px;
        }}
        .stock-code {{
            font-family: "SF Mono", Monaco, monospace;
            padding: 1px 5px;
            border-radius: 4px;
        }}
        .stock-name {{
            opacity: 0.85;
        }}
        .ann-title {{
            font-size: 11.5px;
            line-height: 1.4;
            color: #475569;
            word-break: break-word;
            display: -webkit-box;
            -webkit-line-clamp: 2;
            -webkit-box-orient: vertical;
            overflow: hidden;
        }}
        .empty {{
            flex: 1;
            display: flex;
            align-items: center;
            justify-content: center;
            color: #e2e8f0;
            font-size: 12px;
        }}
        footer {{
            text-align: center;
            margin-top: 20px;
            padding-top: 14px;
            border-top: 1px solid #e2e8f0;
            font-size: 12px;
            color: #94a3b8;
        }}
        /* 侧边栏 */
        .sidebar-overlay {{
            position: fixed;
            inset: 0;
            background: rgba(15, 23, 42, 0.35);
            backdrop-filter: blur(2px);
            opacity: 0;
            visibility: hidden;
            transition: opacity 0.25s ease, visibility 0.25s ease;
            z-index: 998;
        }}
        .sidebar-overlay.active {{
            opacity: 1;
            visibility: visible;
        }}
        .sidebar {{
            position: fixed;
            top: 0;
            right: 0;
            width: 720px;
            max-width: 92vw;
            height: 100vh;
            background: #fff;
            box-shadow: -4px 0 24px rgba(0,0,0,0.12);
            transform: translateX(100%);
            transition: transform 0.3s cubic-bezier(0.4, 0, 0.2, 1);
            z-index: 999;
            display: flex;
            flex-direction: column;
        }}
        .sidebar.active {{
            transform: translateX(0);
        }}
        .sidebar-header {{
            display: flex;
            align-items: center;
            justify-content: space-between;
            padding: 14px 18px;
            border-bottom: 1px solid #e2e8f0;
            flex-shrink: 0;
        }}
        .sidebar-title {{
            font-size: 14px;
            font-weight: 600;
            color: #0f172a;
        }}
        .sidebar-close {{
            width: 32px;
            height: 32px;
            border: none;
            background: #f1f5f9;
            border-radius: 8px;
            cursor: pointer;
            font-size: 18px;
            color: #64748b;
            display: flex;
            align-items: center;
            justify-content: center;
            transition: background 0.15s;
        }}
        .sidebar-close:hover {{
            background: #e2e8f0;
            color: #0f172a;
        }}
        .sidebar-body {{
            flex: 1;
            overflow: hidden;
            position: relative;
        }}
        .sidebar-iframe {{
            width: 100%;
            height: 100%;
            border: none;
        }}
        /* 密码锁屏 */
        .auth-overlay {{
            position: fixed;
            inset: 0;
            background: linear-gradient(135deg, #0f172a 0%, #1e293b 100%);
            z-index: 2000;
            display: flex;
            align-items: center;
            justify-content: center;
            transition: opacity 0.4s ease, visibility 0.4s ease;
        }}
        .auth-overlay.hidden {{
            opacity: 0;
            visibility: hidden;
            pointer-events: none;
        }}
        .auth-card {{
            background: #fff;
            border-radius: 16px;
            padding: 36px 32px;
            width: 360px;
            max-width: 90vw;
            box-shadow: 0 20px 60px rgba(0,0,0,0.3);
            text-align: center;
        }}
        .auth-icon {{
            font-size: 42px;
            margin-bottom: 12px;
        }}
        .auth-title {{
            font-size: 20px;
            font-weight: 700;
            color: #0f172a;
            margin-bottom: 6px;
        }}
        .auth-subtitle {{
            font-size: 13px;
            color: #64748b;
            margin-bottom: 24px;
        }}
        .auth-input {{
            width: 100%;
            padding: 12px 14px;
            border: 1px solid #e2e8f0;
            border-radius: 10px;
            font-size: 15px;
            color: #0f172a;
            outline: none;
            transition: border-color 0.15s, box-shadow 0.15s;
            margin-bottom: 12px;
        }}
        .auth-input:focus {{
            border-color: #3b82f6;
            box-shadow: 0 0 0 3px rgba(59,130,246,0.1);
        }}
        .auth-btn {{
            width: 100%;
            padding: 12px;
            background: #3b82f6;
            color: #fff;
            border: none;
            border-radius: 10px;
            font-size: 15px;
            font-weight: 600;
            cursor: pointer;
            transition: background 0.15s;
        }}
        .auth-btn:hover {{
            background: #2563eb;
        }}
        .auth-error {{
            font-size: 13px;
            color: #ef4444;
            margin-top: 10px;
            min-height: 20px;
        }}
        /* 响应式 */
        @media (max-width: 768px) {{
            .calendar-grid {{
                grid-template-columns: repeat(7, 1fr);
                grid-template-rows: repeat(6, minmax(90px, auto));
                gap: 4px;
            }}
            .day-cell {{
                min-height: 90px;
            }}
            .day-body {{
                max-height: 120px;
            }}
            .ann-title {{
                -webkit-line-clamp: 1;
                font-size: 10px;
            }}
            header h1 {{
                font-size: 20px;
            }}
            .control-bar, .filter-bar {{
                padding: 10px 12px;
            }}
        }}
    </style>
</head>
<body>
    <!-- 密码锁屏层 -->
    <div class="auth-overlay" id="authOverlay">
        <div class="auth-card">
            <div class="auth-icon">🔒</div>
            <div class="auth-title">港股公告日历</div>
            <div class="auth-subtitle">请输入账号和密码</div>
            <input type="text" class="auth-input" id="authUser" placeholder="账号" onkeydown="if(event.key==='Enter')checkAuth()" style="margin-bottom:8px;">
            <input type="password" class="auth-input" id="authPass" placeholder="密码" onkeydown="if(event.key==='Enter')checkAuth()">
            <button class="auth-btn" onclick="checkAuth()">进入</button>
            <div class="auth-error" id="authError"></div>
        </div>
    </div>

    <div class="container">
        <header>
            <h1>📅 港股公告日历</h1>
            <div style="display:flex;gap:12px;align-items:center;">
                <button class="holdings-btn" onclick="openHoldingsModal()">📋 持仓管理</button>
                <div class="update-time">更新时间：{UPDATE_TIME}</div>
            </div>
        </header>

        <!-- 年月选择 -->
        <div class="control-bar">
            <button class="today-btn" onclick="prevMonth()" style="background:#64748b; padding:7px 12px;">◀</button>
            <div class="control-group">
                <span class="control-label">年份</span>
                <select class="control-select" id="yearSelect" onchange="onYearMonthChange()"></select>
            </div>
            <div class="control-group">
                <span class="control-label">月份</span>
                <select class="control-select" id="monthSelect" onchange="onYearMonthChange()"></select>
            </div>
            <button class="today-btn" onclick="goToday()">今天</button>
            <button class="today-btn" onclick="nextMonth()" style="background:#64748b; padding:7px 12px;">▶</button>
        </div>

        <!-- 基金筛选 -->
        <div class="filter-section">
            <div class="filter-bar" id="fundFilterBar">
                <span class="filter-label">基金/账户：</span>
                <!-- JS 填充 -->
            </div>
            <div class="filter-bar" id="stockFilterBar">
                <span class="filter-label">显示股票：</span>
                <!-- JS 填充 -->
            </div>
            <div class="filter-bar" id="categoryFilterBar">
                <span class="filter-label">公告类型：</span>
                <!-- JS 填充 -->
            </div>
        </div>

        <!-- 星期标题 -->
        <div class="weekday-header">
            <div class="weekday-cell">日</div>
            <div class="weekday-cell">一</div>
            <div class="weekday-cell">二</div>
            <div class="weekday-cell">三</div>
            <div class="weekday-cell">四</div>
            <div class="weekday-cell">五</div>
            <div class="weekday-cell">六</div>
        </div>

        <!-- 日历网格 -->
        <div class="calendar-grid" id="calendarGrid"></div>

        <footer>
            数据来源：东方财富 · 港交所披露易同步
        </footer>
    </div>

    <!-- 持仓管理模态框 -->
    <div class="holdings-overlay" id="holdingsOverlay" onclick="closeHoldingsModal(event)">
        <div class="holdings-modal" onclick="event.stopPropagation()">
            <div class="holdings-header">
                <span class="holdings-title">📋 持仓管理</span>
                <button class="holdings-close" onclick="closeHoldingsModal()">×</button>
            </div>
            <div class="holdings-body">
                <div class="holdings-tab-bar">
                    <button class="holdings-tab active" id="tabCurrent" onclick="switchHoldingsTab('current')">当前持仓</button>
                    <button class="holdings-tab" id="tabUpload" onclick="switchHoldingsTab('upload')">上传更新</button>
                </div>
                <div id="holdingsCurrentPanel">
                    <div id="holdingsList"></div>
                </div>
                <div id="holdingsUploadPanel" style="display:none;">
                    <div class="holdings-upload-area" id="uploadArea" onclick="document.getElementById('fileInput').click()">
                        <div style="font-size:28px;margin-bottom:8px;">📁</div>
                        <div style="font-size:14px;font-weight:500;color:#334155;margin-bottom:4px;">点击选择 Excel 文件</div>
                        <div style="font-size:12px;">支持 .xlsx 格式，第2个sheet为持仓汇总</div>
                    </div>
                    <input type="file" id="fileInput" accept=".xlsx,.xls" style="display:none;" onchange="handleFileUpload(this)">
                    <div class="holdings-preview" id="uploadPreview" style="display:none;"></div>
                </div>
            </div>
            <div class="holdings-actions">
                <button class="btn-primary" onclick="downloadHoldings()">⬇️ 下载当前持仓 JSON</button>
                <button class="btn-secondary" id="downloadNewBtn" style="display:none;" onclick="downloadNewHoldings()">⬇️ 下载新持仓 JSON</button>
                <div style="margin-left:auto;font-size:12px;color:#94a3b8;" id="holdingsCount"></div>
            </div>
        </div>
    </div>

    <!-- 侧边栏 -->
    <div class="sidebar-overlay" id="sidebarOverlay" onclick="closeSidebar()"></div>
    <div class="sidebar" id="sidebar">
        <div class="sidebar-header">
            <span class="sidebar-title">公告详情</span>
            <button class="sidebar-close" onclick="closeSidebar()">×</button>
        </div>
        <div class="sidebar-body">
            <iframe class="sidebar-iframe" id="sidebarIframe" src=""></iframe>
        </div>
    </div>

    <script>
        // 密码验证
        const AUTH_USER = 'tfiam';
        const AUTH_PASSWORD = 'tfi2026';
        (function() {
            const overlay = document.getElementById('authOverlay');
            const userInput = document.getElementById('authUser');
            const passInput = document.getElementById('authPass');
            const error = document.getElementById('authError');
            
            // 如果已验证过，直接跳过
            if (localStorage.getItem('hkcalendar_auth') === '1') {
                overlay.classList.add('hidden');
                return;
            }
            
            // 自动聚焦账号输入框
            setTimeout(() => userInput.focus(), 100);
            
            window.checkAuth = function() {
                const user = userInput.value.trim();
                const pass = passInput.value.trim();
                console.log('尝试登录:', '账号长度=' + user.length, '密码长度=' + pass.length);
                if (user === AUTH_USER && pass === AUTH_PASSWORD) {
                    try {
                        localStorage.setItem('hkcalendar_auth', '1');
                    } catch (e) {
                        console.warn('localStorage 不可用:', e);
                    }
                    overlay.classList.add('hidden');
                    error.textContent = '';
                } else {
                    error.textContent = '账号或密码错误（账号: tfiam / 密码: tfi2026）';
                    passInput.value = '';
                    passInput.focus();
                }
            };
        })();

        // 嵌入数据
        const annData = {DATA_JSON};
        const stockList = {STOCKS_JSON};
        const fundList = {FUNDS_JSON};
        const categoryList = {CATEGORIES_JSON};
        const fundColorMap = {{}};
        fundList.forEach(f => fundColorMap[f.name] = f.color);
        const todayStr = "{TODAY_STR}";

        let currentYear = new Date().getFullYear();
        let currentMonth = new Date().getMonth();
        let selectedFunds = new Set(fundList.map(f => f.name));
        let selectedCodes = new Set(stockList.map(s => s.code));
        let selectedCategories = new Set(categoryList.map(c => c.name));

        const monthNames = ["1月", "2月", "3月", "4月", "5月", "6月", "7月", "8月", "9月", "10月", "11月", "12月"];

        // 初始化年月下拉框
        function initSelectors() {{
            const yearSel = document.getElementById("yearSelect");
            const monthSel = document.getElementById("monthSelect");
            const now = new Date();
            const startYear = now.getFullYear() - 3;
            const endYear = now.getFullYear() + 3;
            for (let y = startYear; y <= endYear; y++) {{
                const opt = document.createElement("option");
                opt.value = y;
                opt.textContent = y + "年";
                if (y === currentYear) opt.selected = true;
                yearSel.appendChild(opt);
            }}
            for (let m = 0; m < 12; m++) {{
                const opt = document.createElement("option");
                opt.value = m;
                opt.textContent = monthNames[m];
                if (m === currentMonth) opt.selected = true;
                monthSel.appendChild(opt);
            }}
        }}

        // 初始化基金筛选栏
        function initFundFilter() {{
            const bar = document.getElementById("fundFilterBar");
            fundList.forEach(f => {{
                const item = document.createElement("label");
                item.className = "filter-item";
                item.innerHTML = `
                    <input type="checkbox" checked value="${{f.name}}" onchange="onFundFilterChange()">
                    <span class="filter-dot" style="background-color: ${{f.color}};"></span>
                    <span class="filter-text">${{f.name}}</span>
                `;
                bar.appendChild(item);
            }});
            const actions = document.createElement("div");
            actions.className = "filter-actions";
            actions.innerHTML = `
                <button class="filter-action-btn" onclick="selectAllFunds(true)">全选</button>
                <button class="filter-action-btn" onclick="selectAllFunds(false)">全不选</button>
            `;
            bar.appendChild(actions);
        }}

        // 初始化股票筛选栏
        function initStockFilter() {{
            const bar = document.getElementById("stockFilterBar");
            // 清空现有内容（保留标签）
            while (bar.children.length > 1) bar.removeChild(bar.lastChild);

            const visibleStocks = stockList.filter(s => selectedFunds.has(s.fund));
            visibleStocks.sort((a, b) => parseInt(a.code) - parseInt(b.code));
            visibleStocks.forEach(s => {
                const item = document.createElement("label");
                item.className = "filter-item";
                const isChecked = selectedCodes.has(s.code) ? "checked" : "";
                item.innerHTML = `
                    <input type="checkbox" ${{isChecked}} value="${{s.code}}" onchange="onStockFilterChange()">
                    <span class="filter-dot" style="background-color: ${{s.color}};"></span>
                    <span class="filter-text"><span class="stock-code">${{s.code}}</span><span class="stock-name">${{s.name_cn}}</span></span>
                `;
                bar.appendChild(item);
            }});
            const actions = document.createElement("div");
            actions.className = "filter-actions";
            actions.innerHTML = `
                <button class="filter-action-btn" onclick="selectAllStocks(true)">全选</button>
                <button class="filter-action-btn" onclick="selectAllStocks(false)">全不选</button>
            `;
            bar.appendChild(actions);
        }}

        // 初始化公告类型筛选栏
        function initCategoryFilter() {{
            const bar = document.getElementById("categoryFilterBar");
            // 清空现有内容（保留标签）
            while (bar.children.length > 1) bar.removeChild(bar.lastChild);

            categoryList.forEach(c => {{
                const item = document.createElement("label");
                item.className = "filter-item";
                item.innerHTML = `
                    <input type="checkbox" checked value="${{c.name}}" onchange="onCategoryFilterChange()">
                    <span class="filter-dot" style="background-color: ${{c.color}};"></span>
                    <span class="filter-text">${{c.name}}</span>
                `;
                bar.appendChild(item);
            }});
            const actions = document.createElement("div");
            actions.className = "filter-actions";
            actions.innerHTML = `
                <button class="filter-action-btn" onclick="selectAllCategories(true)">全选</button>
                <button class="filter-action-btn" onclick="selectAllCategories(false)">全不选</button>
            `;
            bar.appendChild(actions);
        }}

        function onFundFilterChange() {{
            const checkboxes = document.querySelectorAll('#fundFilterBar input[type="checkbox"]');
            selectedFunds = new Set();
            checkboxes.forEach(cb => {{
                if (cb.checked) selectedFunds.add(cb.value);
            }});
            // 更新股票筛选栏（只显示选中基金下的股票）
            initStockFilter();
            renderCalendar(currentYear, currentMonth);
        }}

        function onStockFilterChange() {{
            const checkboxes = document.querySelectorAll('#stockFilterBar input[type="checkbox"]');
            selectedCodes = new Set();
            checkboxes.forEach(cb => {{
                if (cb.checked) selectedCodes.add(cb.value);
            }});
            renderCalendar(currentYear, currentMonth);
        }}

        function onCategoryFilterChange() {{
            const checkboxes = document.querySelectorAll('#categoryFilterBar input[type="checkbox"]');
            selectedCategories = new Set();
            checkboxes.forEach(cb => {{
                if (cb.checked) selectedCategories.add(cb.value);
            }});
            renderCalendar(currentYear, currentMonth);
        }}

        function selectAllFunds(checked) {{
            const checkboxes = document.querySelectorAll('#fundFilterBar input[type="checkbox"]');
            checkboxes.forEach(cb => cb.checked = checked);
            onFundFilterChange();
        }}

        function selectAllStocks(checked) {{
            const checkboxes = document.querySelectorAll('#stockFilterBar input[type="checkbox"]');
            checkboxes.forEach(cb => cb.checked = checked);
            onStockFilterChange();
        }}

        function selectAllCategories(checked) {{
            const checkboxes = document.querySelectorAll('#categoryFilterBar input[type="checkbox"]');
            checkboxes.forEach(cb => cb.checked = checked);
            onCategoryFilterChange();
        }}

        function onYearMonthChange() {{
            currentYear = parseInt(document.getElementById("yearSelect").value);
            currentMonth = parseInt(document.getElementById("monthSelect").value);
            renderCalendar(currentYear, currentMonth);
        }}

        function goToday() {{
            const now = new Date();
            currentYear = now.getFullYear();
            currentMonth = now.getMonth();
            document.getElementById("yearSelect").value = currentYear;
            document.getElementById("monthSelect").value = currentMonth;
            renderCalendar(currentYear, currentMonth);
        }}

        function prevMonth() {{
            currentMonth--;
            if (currentMonth < 0) {{
                currentMonth = 11;
                currentYear--;
            }}
            document.getElementById("yearSelect").value = currentYear;
            document.getElementById("monthSelect").value = currentMonth;
            renderCalendar(currentYear, currentMonth);
        }}

        function nextMonth() {{
            currentMonth++;
            if (currentMonth > 11) {{
                currentMonth = 0;
                currentYear++;
            }}
            document.getElementById("yearSelect").value = currentYear;
            document.getElementById("monthSelect").value = currentMonth;
            renderCalendar(currentYear, currentMonth);
        }}

        function renderCalendar(year, month) {{
            const grid = document.getElementById("calendarGrid");
            grid.innerHTML = "";

            const firstDay = new Date(year, month, 1);
            const startDayOfWeek = firstDay.getDay();
            const daysInMonth = new Date(year, month + 1, 0).getDate();
            const prevMonthDays = new Date(year, month, 0).getDate();
            const totalCells = 42;

            for (let i = 0; i < totalCells; i++) {{
                let dayNum, isCurrentMonth, cellMonth, cellYear;

                if (i < startDayOfWeek) {{
                    dayNum = prevMonthDays - startDayOfWeek + 1 + i;
                    const d = new Date(year, month, 1 - startDayOfWeek + i);
                    cellYear = d.getFullYear();
                    cellMonth = d.getMonth();
                    isCurrentMonth = false;
                }} else if (i < startDayOfWeek + daysInMonth) {{
                    dayNum = i - startDayOfWeek + 1;
                    cellYear = year;
                    cellMonth = month;
                    isCurrentMonth = true;
                }} else {{
                    dayNum = i - startDayOfWeek - daysInMonth + 1;
                    const d = new Date(year, month, daysInMonth + (i - startDayOfWeek - daysInMonth + 1));
                    cellYear = d.getFullYear();
                    cellMonth = d.getMonth();
                    isCurrentMonth = false;
                }}

                const mm = String(cellMonth + 1).padStart(2, "0");
                const dd = String(dayNum).padStart(2, "0");
                const dateStr = `${{cellYear}}-${{mm}}-${{dd}}`;
                const isToday = dateStr === todayStr;

                // 过滤：选中的基金 AND 选中的股票 AND 选中的公告类型（有交集即显示）
                let anns = (annData[dateStr] || []).filter(a => {
                    const fundMatch = selectedFunds.has(a.fund);
                    const codeMatch = selectedCodes.has(a.code);
                    const catMatch = a.categories && a.categories.some(c => selectedCategories.has(c));
                    return fundMatch && codeMatch && catMatch;
                });

                const cell = document.createElement("div");
                cell.className = "day-cell" + (isToday ? " today" : "") + (isCurrentMonth ? "" : " other-month");

                let cardsHtml = "";
                if (anns.length > 0) {{
                    cardsHtml = anns.map(a => `
                        <div class="ann-card" title="${{a.title}}" style="border-left-color: ${{a.fundColor}};" onclick="if('${{a.pdf_url}}')openSidebar('${{a.pdf_url}}')">
                            <div class="fund-tag" style="background-color: ${{a.fundColor}}15; color: ${{a.fundColor}};">${{a.fund}}</div>
                            <div class="stock-tag">
                                <span class="stock-code" style="background-color: ${{a.color}}20; color: ${{a.color}}; border: 1px solid ${{a.color}}40; border-radius: 4px;">${{a.code}}</span>
                                <span class="stock-name" style="color: ${{a.color}}; font-size:10px;">${{a.name_cn}}</span>
                            </div>
                            <div class="ann-title">${{a.title}} <span style="color:#94a3b8;font-size:10px;">↗</span></div>
                        </div>
                    `).join("");
                }} else {{
                    cardsHtml = '<div class="empty"></div>';
                }}

                cell.innerHTML = `
                    <div class="day-header-cell">
                        <span class="day-number">${{dayNum}}</span>
                        ${{(anns.length > 0 && isCurrentMonth) ? `<span class="ann-count">${{anns.length}}</span>` : ""}
                    </div>
                    <div class="day-body">${{cardsHtml}}</div>
                `;

                grid.appendChild(cell);
            }}
        }}

        // 侧边栏
        function openSidebar(url) {{
            const sidebar = document.getElementById('sidebar');
            const overlay = document.getElementById('sidebarOverlay');
            const iframe = document.getElementById('sidebarIframe');
            if (!url) return;
            iframe.src = url;
            sidebar.classList.add('active');
            overlay.classList.add('active');
            document.body.style.overflow = 'hidden';
        }}
        function closeSidebar() {{
            const sidebar = document.getElementById('sidebar');
            const overlay = document.getElementById('sidebarOverlay');
            const iframe = document.getElementById('sidebarIframe');
            sidebar.classList.remove('active');
            overlay.classList.remove('active');
            document.body.style.overflow = '';
            setTimeout(() => {{ iframe.src = ''; }}, 300);
        }}
        document.addEventListener('keydown', e => {{
            if (e.key === 'Escape') closeSidebar();
        }});

        // 持仓管理
        function openHoldingsModal() {{
            document.getElementById('holdingsOverlay').classList.add('active');
            renderHoldingsList();
        }}
        function closeHoldingsModal(e) {{
            if (e && e.target !== document.getElementById('holdingsOverlay')) return;
            document.getElementById('holdingsOverlay').classList.remove('active');
        }}
        function switchHoldingsTab(tab) {{
            document.getElementById('tabCurrent').classList.toggle('active', tab === 'current');
            document.getElementById('tabUpload').classList.toggle('active', tab === 'upload');
            document.getElementById('holdingsCurrentPanel').style.display = tab === 'current' ? 'block' : 'none';
            document.getElementById('holdingsUploadPanel').style.display = tab === 'upload' ? 'block' : 'none';
            document.getElementById('downloadNewBtn').style.display = tab === 'upload' && uploadedHoldings ? 'block' : 'none';
        }}
        function renderHoldingsList() {{
            const container = document.getElementById('holdingsList');
            const groups = {{}};
            stockList.forEach(s => {{
                if (!groups[s.fund]) groups[s.fund] = [];
                groups[s.fund].push(s);
            }});
            let html = '';
            let total = 0;
            for (const [fund, stocks] of Object.entries(groups)) {{
                const fundColor = fundColorMap[fund] || '#64748b';
                html += `<div class="holdings-group">
                    <div class="holdings-group-title">
                        <span class="holdings-group-dot" style="background:${{fundColor}}"></span>
                        ${{fund}} (${{stocks.length}}只)
                    </div>
                    <table class="holdings-table">
                        <thead><tr><th>代码</th><th>名称</th><th>账户</th></tr></thead>
                        <tbody>`;
                stocks.forEach(s => {{
                    html += `<tr><td style="font-family:monospace;font-weight:600;">${{s.code}}</td><td>${{s.name_cn || s.name}}</td><td style="color:#64748b;">${{s.account || '-'}}</td></tr>`;
                }});
                html += '</tbody></table></div>';
                total += stocks.length;
            }}
            container.innerHTML = html || '<div style="color:#94a3b8;text-align:center;padding:20px;">暂无持仓数据</div>';
            document.getElementById('holdingsCount').textContent = `共 ${{total}} 只`;
        }}
        function downloadHoldings() {{
            const data = stockList.map(s => ({{
                code: s.code,
                name: s.name,
                name_cn: s.name_cn,
                fund: s.fund,
                account: s.account || ''
            }}));
            const blob = new Blob([JSON.stringify(data, null, 2)], {{type: 'application/json'}});
            const url = URL.createObjectURL(blob);
            const a = document.createElement('a');
            a.href = url;
            a.download = 'holdings_' + new Date().toISOString().slice(0,10) + '.json';
            a.click();
            URL.revokeObjectURL(url);
        }}
        let uploadedHoldings = null;
        async function handleFileUpload(input) {{
            const file = input.files[0];
            if (!file) return;
            const area = document.getElementById('uploadArea');
            area.innerHTML = '<div style="color:#3b82f6;font-weight:500;">正在解析...</div>';
            try {{
                // 动态加载 SheetJS
                if (typeof XLSX === 'undefined') {{
                    await new Promise((resolve, reject) => {{
                        const script = document.createElement('script');
                        script.src = 'https://cdn.sheetjs.com/xlsx-0.20.2/package/dist/xlsx.full.min.js';
                        script.onload = resolve;
                        script.onerror = reject;
                        document.head.appendChild(script);
                    }});
                }}
                const data = await file.arrayBuffer();
                const workbook = XLSX.read(data, {{type: 'array'}});
                // 使用第二个 sheet（持仓汇总）
                const sheetName = workbook.SheetNames[1] || workbook.SheetNames[0];
                const worksheet = workbook.Sheets[sheetName];
                const json = XLSX.utils.sheet_to_json(worksheet, {{header: 1}});
                if (json.length < 2) throw new Error('表格数据为空');
                // 解析表头，找到对应列
                const headers = json[0].map(h => String(h).trim());
                const colIndex = {{}};
                headers.forEach((h, i) => {{
                    if (h.includes('代码') || h === 'code') colIndex.code = i;
                    if (h.includes('名称') || h === 'name') colIndex.name = i;
                    if (h.includes('基金') || h === 'fund') colIndex.fund = i;
                    if (h.includes('账户') || h === 'account') colIndex.account = i;
                }});
                // 如果没找到，按固定位置解析（TFI 格式: 代码,名称,基金,账户,...）
                const codeIdx = colIndex.code ?? 0;
                const nameIdx = colIndex.name ?? 1;
                const fundIdx = colIndex.fund ?? 2;
                const accountIdx = colIndex.account ?? 3;
                const results = [];
                const seen = new Set();
                for (let i = 1; i < json.length; i++) {{
                    const row = json[i];
                    if (!row || !row[codeIdx]) continue;
                    const code = String(row[codeIdx]).trim().replace(/\.0$/, '');
                    if (!code || seen.has(code)) continue;
                    seen.add(code);
                    const name = String(row[nameIdx] || '').trim();
                    const fund = String(row[fundIdx] || '').trim();
                    const account = String(row[accountIdx] || '').trim();
                    const cnMatch = name.match(/^[\u4e00-\u9fa5]+/);
                    results.push({{code, name, name_cn: cnMatch ? cnMatch[0] : name, fund, account}});
                }}
                uploadedHoldings = results;
                // 显示预览
                let previewHtml = `<div style="font-weight:600;margin-bottom:8px;color:#0f172a;">解析成功：${{results.length}} 只股票</div>`;
                previewHtml += '<table class="holdings-table"><thead><tr><th>代码</th><th>名称</th><th>基金</th><th>账户</th></tr></thead><tbody>';
                results.slice(0, 20).forEach(r => {{
                    previewHtml += `<tr><td style="font-family:monospace;font-weight:600;">${{r.code}}</td><td>${{r.name_cn}}</td><td>${{r.fund}}</td><td>${{r.account || '-'}}</td></tr>`;
                }});
                if (results.length > 20) previewHtml += `<tr><td colspan="4" style="text-align:center;color:#94a3b8;">... 还有 ${{results.length - 20}} 只 ...</td></tr>`;
                previewHtml += '</tbody></table>';
                document.getElementById('uploadPreview').innerHTML = previewHtml;
                document.getElementById('uploadPreview').style.display = 'block';
                document.getElementById('downloadNewBtn').style.display = 'block';
                area.innerHTML = `<div style="color:#10b981;font-weight:500;">✅ 解析成功：${{results.length}} 只股票</div><div style="font-size:12px;color:#64748b;margin-top:4px;">${{file.name}}</div>`;
            }} catch (err) {{
                area.innerHTML = `<div style="color:#ef4444;font-weight:500;">❌ 解析失败</div><div style="font-size:12px;color:#64748b;margin-top:4px;">${{err.message}}</div>`;
                console.error(err);
            }}
        }}
        function downloadNewHoldings() {{
            if (!uploadedHoldings) return;
            const blob = new Blob([JSON.stringify(uploadedHoldings, null, 2)], {{type: 'application/json'}});
            const url = URL.createObjectURL(blob);
            const a = document.createElement('a');
            a.href = url;
            a.download = 'holdings_new_' + new Date().toISOString().slice(0,10) + '.json';
            a.click();
            URL.revokeObjectURL(url);
        }}

        // 初始化
        initSelectors();
        initFundFilter();
        initStockFilter();
        initCategoryFilter();
        renderCalendar(currentYear, currentMonth);
    </script>
</body>
</html>'''

    return (template
        .replace('{DATA_JSON}', data_json)
        .replace('{STOCKS_JSON}', stocks_json)
        .replace('{FUNDS_JSON}', funds_json)
        .replace('{CATEGORIES_JSON}', categories_json)
        .replace('{TODAY_STR}', today_str)
        .replace('{UPDATE_TIME}', update_time)
        .replace('{{', '{')
        .replace('}}', '}')
    )


if __name__ == "__main__":
    main()
